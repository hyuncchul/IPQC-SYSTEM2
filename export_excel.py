import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
from datetime import datetime

def thin_border():
    thin = Side(style='thin')
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def header_fill(color='1F4E79'):
    return PatternFill('solid', start_color=color, end_color=color)

def export_daily_excel(date_str, all_data):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ===== 停機&尾批交接 sheet =====
    ws_ho = wb.create_sheet('停機&尾批交接')
    _write_handover_sheet(ws_ho, date_str, all_data['handovers'])

    # ===== Per machine sheets =====
    for item in all_data['entries']:
        entry = item['entry']
        machine_id = entry['machine_id']
        shift_name = {'morning': '早班', 'night': '夜班', 'first_piece': '首件'}.get(entry['shift'], entry['shift'])
        sheet_name = f"{machine_id}_{shift_name}"[:31]
        ws = wb.create_sheet(sheet_name)
        _write_machine_sheet(ws, entry, item['visual'], item['eol'], item['dims'], date_str)

    # ===== 異常記錄 sheet =====
    if all_data['abnormalities']:
        ws_ab = wb.create_sheet('異常記錄')
        _write_abnormality_sheet(ws_ab, date_str, all_data['abnormalities'])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def _style_header_cell(cell, text, bg='1F4E79', fg='FFFFFF', bold=True):
    cell.value = text
    cell.font = Font(bold=bold, color=fg, name='Arial', size=10)
    cell.fill = PatternFill('solid', start_color=bg, end_color=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border()

def _style_data_cell(cell, value='', bg=None, bold=False):
    cell.value = value
    cell.font = Font(name='Arial', size=10, bold=bold)
    if bg:
        cell.fill = PatternFill('solid', start_color=bg, end_color=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border()

def _write_machine_sheet(ws, entry, visual, eol, dims, date_str):
    machine_id = entry['machine_id']
    shift_label = {'morning': '早班', 'night': '夜班', 'first_piece': '首件'}.get(entry['shift'], entry['shift'])

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18

    # Title
    ws.merge_cells('A1:E1')
    title = ws['A1']
    title.value = f'{machine_id} QC 檢查記錄 - {date_str} {shift_label}'
    title.font = Font(bold=True, size=13, name='Arial', color='FFFFFF')
    title.fill = PatternFill('solid', start_color='1F4E79', end_color='1F4E79')
    title.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Info row
    row = 2
    info_labels = ['機台', 'Part No.', 'Lot No.', '作業者', '備註']
    info_vals = [machine_id, entry.get('part_no',''), entry.get('lot_no',''),
                 entry.get('submitted_by',''), entry.get('notes','')]
    for col, (lbl, val) in enumerate(zip(info_labels, info_vals), 1):
        c = ws.cell(row=row, column=col)
        c.value = f'{lbl}: {val}'
        c.font = Font(name='Arial', size=9)
        c.border = thin_border()
        c.alignment = Alignment(wrap_text=True)

    # Visual section
    row = 4
    ws.merge_cells(f'A{row}:E{row}')
    _style_header_cell(ws.cell(row=row, column=1), '外觀注意事項', bg='2E75B6')
    row += 1
    for lbl, col in [('No.','A'),('檢查項目','B'),('SPEC','C'),('結果','D'),('拒收Lot#','E')]:
        _style_header_cell(ws[f'{col}{row}'], lbl, bg='4472C4')
    row += 1
    for i, v in enumerate(visual):
        ws.cell(row=row, column=1).value = i+1
        ws.cell(row=row, column=2).value = v.get('item_name','')
        ws.cell(row=row, column=3).value = ''
        res = v.get('result','')
        c = ws.cell(row=row, column=4)
        c.value = res
        if res and res.upper() in ['NG','X','拒收']:
            c.fill = PatternFill('solid', start_color='FF0000', end_color='FF0000')
            c.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        elif res and res.upper() in ['OK','O','PASS']:
            c.fill = PatternFill('solid', start_color='70AD47', end_color='70AD47')
            c.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        else:
            _style_data_cell(c, res)
        ws.cell(row=row, column=5).value = v.get('rejected_lot','')
        for col in range(1,6):
            ws.cell(row=row, column=col).border = thin_border()
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[row].height = 20
        row += 1

    # EOL section
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    _style_header_cell(ws.cell(row=row, column=1), 'EOL 檢查', bg='2E75B6')
    row += 1
    for lbl, col in [('No.','A'),('檢查項目','B'),('SPEC','C'),('結果','D'),('拒收Lot#','E')]:
        _style_header_cell(ws[f'{col}{row}'], lbl, bg='4472C4')
    row += 1
    for i, v in enumerate(eol):
        ws.cell(row=row, column=1).value = i+1
        ws.cell(row=row, column=2).value = v.get('item_name','')
        ws.cell(row=row, column=3).value = ''
        ws.cell(row=row, column=4).value = v.get('result','')
        ws.cell(row=row, column=5).value = v.get('rejected_lot','')
        for col in range(1,6):
            ws.cell(row=row, column=col).border = thin_border()
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[row].height = 18
        row += 1

    # Dimension section
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    _style_header_cell(ws.cell(row=row, column=1), '尺寸注意事項', bg='2E75B6')
    row += 1
    for lbl, col in [('No.','A'),('測定項目','B'),('SPEC','C'),('測定值','D'),('備註','E')]:
        _style_header_cell(ws[f'{col}{row}'], lbl, bg='4472C4')
    row += 1
    for i, v in enumerate(dims):
        ws.cell(row=row, column=1).value = i+1
        ws.cell(row=row, column=2).value = v.get('item_name','')
        ws.cell(row=row, column=3).value = ''
        ws.cell(row=row, column=4).value = v.get('result','')
        ws.cell(row=row, column=5).value = ''
        for col in range(1,6):
            ws.cell(row=row, column=col).border = thin_border()
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[row].height = 18
        row += 1

def _write_handover_sheet(ws, date_str, handovers):
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 25

    ws.merge_cells('A1:F1')
    ws['A1'].value = f'停機&尾批交接 - {date_str}'
    ws['A1'].font = Font(bold=True, size=13, color='FFFFFF', name='Arial')
    ws['A1'].fill = PatternFill('solid', start_color='1F4E79', end_color='1F4E79')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    for col, lbl in enumerate(['機台','尾批Lot#','停機原因','機台','尾批Lot#','停機原因'],1):
        c = ws.cell(row=2, column=col)
        _style_header_cell(c, lbl, bg='4472C4')

    handover_map = {h['machine_id']: h for h in handovers}
    from machine_config import MACHINE_LIST
    mid = len(MACHINE_LIST)//2
    left_machines = MACHINE_LIST[:mid]
    right_machines = MACHINE_LIST[mid:]

    for i, (lm, rm) in enumerate(zip(left_machines, right_machines), 3):
        lh = handover_map.get(lm, {})
        rh = handover_map.get(rm, {})
        ws.cell(row=i, column=1).value = lm
        ws.cell(row=i, column=2).value = lh.get('last_batch','')
        ws.cell(row=i, column=3).value = lh.get('reason','')
        ws.cell(row=i, column=4).value = rm
        ws.cell(row=i, column=5).value = rh.get('last_batch','')
        ws.cell(row=i, column=6).value = rh.get('reason','')
        for col in range(1,7):
            ws.cell(row=i, column=col).border = thin_border()
            ws.cell(row=i, column=col).alignment = Alignment(horizontal='center', vertical='center')

def _write_abnormality_sheet(ws, date_str, abnormalities):
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 30

    ws.merge_cells('A1:E1')
    ws['A1'].value = f'異常發生記錄 - {date_str}'
    ws['A1'].font = Font(bold=True, size=13, color='FFFFFF', name='Arial')
    ws['A1'].fill = PatternFill('solid', start_color='C00000', end_color='C00000')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    for col, lbl in enumerate(['機台','班別','異常內容','原因分析','對策/措施'],1):
        _style_header_cell(ws.cell(row=2, column=col), lbl, bg='C00000')

    shift_map = {'morning':'早班','night':'夜班','first_piece':'首件'}
    for i, abn in enumerate(abnormalities, 3):
        ws.cell(row=i, column=1).value = abn.get('machine_id','')
        ws.cell(row=i, column=2).value = shift_map.get(abn.get('shift',''), abn.get('shift',''))
        ws.cell(row=i, column=3).value = abn.get('description','')
        ws.cell(row=i, column=4).value = abn.get('cause','')
        ws.cell(row=i, column=5).value = abn.get('countermeasure','')
        for col in range(1,6):
            ws.cell(row=i, column=col).border = thin_border()
            ws.cell(row=i, column=col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[i].height = 25
